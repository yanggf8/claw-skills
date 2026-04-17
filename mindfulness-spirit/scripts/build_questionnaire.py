#!/usr/bin/env python3
"""Build Ping W. persona questionnaire Excel file."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()

# Style constants
HEADER_FONT = Font(name="Microsoft JhengHei", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(name="Microsoft JhengHei", size=11, bold=True)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
Q_FONT = Font(name="Microsoft JhengHei", size=11)
WRAP = Alignment(wrap_text=True, vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_row(ws, row, is_section=False):
    for cell in ws[row]:
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if is_section:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        else:
            cell.font = Q_FONT


def add_section(ws, row, title):
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    style_row(ws, row, is_section=True)
    return row + 1


def add_q(ws, row, num, question):
    ws.cell(row=row, column=1, value=num)
    ws.cell(row=row, column=2, value=question)
    ws.cell(row=row, column=3, value="")
    ws.row_dimensions[row].height = 45
    style_row(ws, row)
    return row + 1


# ── Sheet 1: 聲音與風格 ──
ws1 = wb.active
ws1.title = "聲音與風格"
ws1.append(["#", "問題", "你的回答"])
style_header(ws1)
ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 55
ws1.column_dimensions["C"].width = 50

r = 2
r = add_q(ws1, r, 1, "如果用三個形容詞描述你的寫作風格，你會說是什麼？")
r = add_q(ws1, r, 2, "你寫文章時，腦中想像的讀者是誰？（例如：某位朋友、某種類型的人）")
r = add_q(ws1, r, 3, "你最常被什麼東西啟發？（自然、經典文本、日常對話、身體經驗……）")
r = add_q(ws1, r, 4, "你喜歡用什麼方式開場？例如：說故事、提問、場景描寫、直接破題？")
r = add_q(ws1, r, 5, "你有哪些口頭禪或常用比喻，是朋友一聽就知道「這是你說的」？")
r = add_q(ws1, r, 6, "你會避開哪些語氣或用詞？（例如：說教感、學術腔、特定詞彙）")

# ── Sheet 2: 思考方式 ──
ws2 = wb.create_sheet("思考方式")
ws2.append(["#", "問題", "你的回答"])
style_header(ws2)
ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 55
ws2.column_dimensions["C"].width = 50

r = 2
r = add_q(ws2, r, 7, "當你遇到一個新科技現象，你通常會怎麼開始思考它？")
r = add_q(ws2, r, 8, "你如何看待不同宗教/靈性傳統之間的關係？")
r = add_q(ws2, r, 9, "你相信「科技」和「靈性」是什麼關係？（對立？互補？鏡子？……）")
r = add_q(ws2, r, 10, "當讀者提出質疑或反對意見，你心裡通常怎麼想？")

# ── Sheet 3: 寫作習慣 ──
ws3 = wb.create_sheet("寫作習慣")
ws3.append(["#", "問題", "你的回答"])
style_header(ws3)
ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 55
ws3.column_dimensions["C"].width = 50

r = 2
r = add_q(ws3, r, 11, "你覺得一篇好文章一定要有什麼元素？")
r = add_q(ws3, r, 12, "你喜歡給答案，還是給問題？還是看情況？")
r = add_q(ws3, r, 13, "你通常會把個人經驗寫進文章嗎？什麼時候會、什麼時候不會？")
r = add_q(ws3, r, 14, "如果要用一句話形容你想帶給讀者的感覺，那句話是？")

# ── Sheet 4: 背景與養分 ──
ws4 = wb.create_sheet("背景與養分")
ws4.append(["#", "問題", "你的回答"])
style_header(ws4)
ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 55
ws4.column_dimensions["C"].width = 50

r = 2
r = add_q(ws4, r, 15, "你跟宗教/靈性傳統的個人淵源是什麼？")
r = add_q(ws4, r, 16, "你日常的實踐是什麼？（冥想、祈禱、瑜珈、書法……）")
r = add_q(ws4, r, 17, "有哪本書、哪位老師、或哪段經歷對你影響最深？")
r = add_q(ws4, r, 18, "你最在意的價值是什麼？（自由、真實、慈悲、簡單……）")

out = "/home/yanggf/a/claw-skills/ping-w-questionnaire.xlsx"
wb.save(out)
print(f"Saved: {out}")
